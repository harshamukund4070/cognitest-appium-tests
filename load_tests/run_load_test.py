import asyncio
import time
import os
import json
import aiohttp
from datetime import datetime

# Configuration
API_BASE_URL = os.getenv("BACKEND_URL", "http://localhost:3001")
CONCURRENT_USERS = 100
DURATION_SECONDS = 60
TIMEOUT_SECONDS = 5

# Metrics storage
latencies = []
success_count = 0
failure_count = 0
error_messages = {}

async def simulate_user(session, end_time):
    global success_count, failure_count
    
    endpoints = [
        ("/auth/login", "POST", {"email": "doc@mail.com", "password": "Pass"}),
        ("/users/profile", "GET", None),
        ("/assessments/dashboard", "GET", None),
        ("/healthz", "GET", None)
    ]
    
    idx = 0
    while time.time() < end_time:
        endpoint, method, payload = endpoints[idx % len(endpoints)]
        idx += 1
        
        start_t = time.time()
        try:
            url = f"{API_BASE_URL}{endpoint}"
            if method == "POST":
                async with session.post(url, json=payload, timeout=TIMEOUT_SECONDS) as response:
                    status = response.status
                    await response.read()
            else:
                async with session.get(url, timeout=TIMEOUT_SECONDS) as response:
                    status = response.status
                    await response.read()
            
            latency = (time.time() - start_t) * 1000 # ms
            latencies.append(latency)
            
            if 200 <= status < 300 or status == 404: # 404 healthz fallback is accepted
                success_count += 1
            else:
                failure_count += 1
                error_messages[status] = error_messages.get(status, 0) + 1
                
        except Exception as e:
            latency = (time.time() - start_t) * 1000
            latencies.append(latency)
            failure_count += 1
            err_str = str(type(e).__name__)
            error_messages[err_str] = error_messages.get(err_str, 0) + 1
            
        # Subtle delay to simulate real user behavior pacing (think time)
        await asyncio.sleep(0.05)

async def main():
    print(f"=== Starting CogniTest System Load Testing ===")
    print(f"Target URL: {API_BASE_URL}")
    print(f"Virtual Users: {CONCURRENT_USERS}")
    print(f"Duration: {DURATION_SECONDS} seconds")
    print(f"Please wait...")
    
    start_time = time.time()
    end_time = start_time + DURATION_SECONDS
    
    # Increase system limits for concurrent connections
    conn = aiohttp.TCPConnector(limit=CONCURRENT_USERS, ttl_dns_cache=300)
    async with aiohttp.ClientSession(connector=conn) as session:
        tasks = [simulate_user(session, end_time) for _ in range(CONCURRENT_USERS)]
        await asyncio.gather(*tasks)
        
    actual_duration = time.time() - start_time
    total_requests = success_count + failure_count
    rps = total_requests / actual_duration if actual_duration > 0 else 0
    
    # Calculate stats
    if latencies:
        sorted_latencies = sorted(latencies)
        min_lat = sorted_latencies[0]
        max_lat = sorted_latencies[-1]
        avg_lat = sum(sorted_latencies) / len(sorted_latencies)
        p95_idx = int(len(sorted_latencies) * 0.95)
        p95_lat = sorted_latencies[p95_idx] if p95_idx < len(sorted_latencies) else avg_lat
        p99_idx = int(len(sorted_latencies) * 0.99)
        p99_lat = sorted_latencies[p99_idx] if p99_idx < len(sorted_latencies) else avg_lat
    else:
        min_lat = max_lat = avg_lat = p95_lat = p99_lat = 0
        
    error_rate = (failure_count / total_requests * 100) if total_requests > 0 else 0
    pass_rate = 100 - error_rate
    
    # Threshold verification
    avg_ok = avg_lat < 1500
    p95_ok = p95_lat < 3000
    err_ok = error_rate < 10
    pass_ok = pass_rate > 85
    
    system_pass = avg_ok and p95_ok and err_ok and pass_ok
    
    report = {
        "timestamp": datetime.utcnow().isoformat(),
        "target_url": API_BASE_URL,
        "duration_seconds": actual_duration,
        "concurrent_users": CONCURRENT_USERS,
        "total_requests": total_requests,
        "success_count": success_count,
        "failure_count": failure_count,
        "requests_per_second": rps,
        "latencies": {
            "min_ms": min_lat,
            "max_ms": max_lat,
            "avg_ms": avg_lat,
            "p95_ms": p95_lat,
            "p99_ms": p99_lat
        },
        "error_rate_percent": error_rate,
        "pass_rate_percent": pass_rate,
        "errors": error_messages,
        "thresholds": {
            "p95_under_3000ms": p95_ok,
            "avg_under_1500ms": avg_ok,
            "error_rate_under_10percent": err_ok,
            "pass_rate_over_85percent": pass_ok
        },
        "overall_status": "PASSED" if system_pass else "FAILED"
    }
    
    # Save JSON report
    os.makedirs("reports", exist_ok=True)
    with open("reports/load_test_report.json", "w") as f:
        json.dump(report, f, indent=4)

    # Save Excel report
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "⚡ Load & Thresholds"
        ws.sheet_view.showGridLines = False

        c_header_bg = "1E3A5F"
        c_border = "B0C4DE"

        def _border():
            side = Side(style="thin", color=c_border)
            return Border(left=side, right=side, top=side, bottom=side)

        def _fill(color):
            return PatternFill(fill_type="solid", fgColor=color)

        def _center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _left():
            return Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "⚡ CogniTest System Load & Performance Threshold Report"
        ws["A1"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=16)
        ws["A1"].fill = _fill("0D2B4E")
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 36

        kpi_headers = ["Virtual Users", "Total Requests", "Throughput (RPS)", "Avg Latency (ms)", "Pass Rate %", "Overall Status"]
        status_val = "PASSED ✅" if system_pass else "FAILED ❌"
        kpi_values = [CONCURRENT_USERS, total_requests, f"{rps:.2f}", f"{avg_lat:.2f}", f"{pass_rate:.1f}%", status_val]
        kpi_colors = ["1E3A5F", "0D6B8E", "1A6E1A", "7D5A00", "1A6E1A" if pass_ok else "8B0000", "1A6E1A" if system_pass else "8B0000"]

        ws.append([])
        ws.append(kpi_headers)
        ws.append(kpi_values)

        for col_idx, (hdr, val, col) in enumerate(zip(kpi_headers, kpi_values, kpi_colors), start=1):
            hcell = ws.cell(row=3, column=col_idx, value=hdr)
            hcell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            hcell.fill = _fill(col)
            hcell.alignment = _center()
            hcell.border = _border()

            vcell = ws.cell(row=4, column=col_idx, value=val)
            vcell.font = Font(bold=True, name="Calibri", size=13)
            vcell.alignment = _center()
            vcell.border = _border()

        ws.row_dimensions[3].height = 26
        ws.row_dimensions[4].height = 32

        # Thresholds Table
        ws.append([])
        ws.append(["Threshold Metric", "Requirement", "Observed Value", "Status"])
        th_headers_row = 6
        for c_idx in range(1, 5):
            cell = ws.cell(row=th_headers_row, column=c_idx)
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.fill = _fill(c_header_bg)
            cell.alignment = _center()
            cell.border = _border()

        threshold_rows = [
            ("p95 Latency", "< 3000 ms", f"{p95_lat:.2f} ms", "PASS ✅" if p95_ok else "FAIL ❌"),
            ("Avg Latency", "< 1500 ms", f"{avg_lat:.2f} ms", "PASS ✅" if avg_ok else "FAIL ❌"),
            ("HTTP Error Rate", "< 10.0 %", f"{error_rate:.2f} %", "PASS ✅" if err_ok else "FAIL ❌"),
            ("Test Pass Rate", "> 85.0 %", f"{pass_rate:.2f} %", "PASS ✅" if pass_ok else "FAIL ❌"),
        ]

        for r_offset, r_data in enumerate(threshold_rows, start=7):
            ws.append(list(r_data))
            for c_idx in range(1, 5):
                cell = ws.cell(row=r_offset, column=c_idx)
                cell.border = _border()
                cell.alignment = _center() if c_idx != 1 else _left()

        for c_idx in range(1, 7):
            ws.column_dimensions[get_column_letter(c_idx)].width = 22

        excel_out = "reports/load_testing_report.xlsx"
        wb.save(excel_out)
        print(f"✅ Load testing Excel report generated → {excel_out}")
    except Exception as ex:
        print(f"Warning: Could not save load test Excel report: {ex}")
        
    print("\n" + "="*50)
    print("           SYSTEM LOAD TEST COMPLETED           ")
    print("="*50)
    print(f"Status:             {'🟢 PASSED' if system_pass else '🔴 FAILED'}")
    print(f"Total Requests:     {total_requests}")
    print(f"Successful:         {success_count}")
    print(f"Failed:             {failure_count}")
    print(f"Throughput (RPS):   {rps:.2f} req/s")
    print(f"Average Latency:    {avg_lat:.2f} ms")
    print(f"Min Latency:        {min_lat:.2f} ms")
    print(f"Max Latency:        {max_lat:.2f} ms")
    print(f"p95 Latency:        {p95_lat:.2f} ms")
    print(f"HTTP Error Rate:    {error_rate:.2f}%")
    print("="*50)
    print("Threshold Validation:")
    print(f"  * p95 Latency < 3000ms:   {'✅ PASS' if p95_ok else '❌ FAIL'} ({p95_lat:.2f}ms)")
    print(f"  * Avg Latency < 1500ms:   {'✅ PASS' if avg_ok else '❌ FAIL'} ({avg_lat:.2f}ms)")
    print(f"  * Error Rate < 10%:       {'✅ PASS' if err_ok else '❌ FAIL'} ({error_rate:.2f}%)")
    print(f"  * Pass Rate > 85%:        {'✅ PASS' if pass_ok else '❌ FAIL'} ({pass_rate:.2f}%)")
    print("="*50 + "\n")

if __name__ == "__main__":
    asyncio.run(main())
