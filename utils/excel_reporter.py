"""
Excel Reporter — generates a rich, analysis-ready Excel workbook
from pytest results for the CogniTest Appium test suite.
"""
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from config.config import EXCEL_REPORT, SCREENSHOTS_DIR, REPORTS_DIR


# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG  = "1E3A5F"   # dark blue
C_HEADER_FG  = "FFFFFF"
C_PASS_BG    = "D6F5D6"   # light green
C_PASS_FG    = "1A6E1A"
C_FAIL_BG    = "FFD6D6"   # light red
C_FAIL_FG    = "8B0000"
C_SKIP_BG    = "FFF3CD"   # light amber
C_SKIP_FG    = "7D5A00"
C_TITLE_BG   = "0D2B4E"
C_ALT_ROW    = "EEF4FB"
C_BORDER     = "B0C4DE"

def _border():
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _header_font():
    return Font(bold=True, color=C_HEADER_FG, name="Calibri", size=11)

def _title_font(size=14):
    return Font(bold=True, color=C_HEADER_FG, name="Calibri", size=size)

def _cell_font(bold=False):
    return Font(bold=bold, name="Calibri", size=10)

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


class ExcelReporter:
    """
    Collects test results via `record()` and writes a multi-sheet Excel
    workbook at the end via `save()`.
    """

    def __init__(self):
        self.results = []   # list of dicts
        self.start_time = datetime.now()
        os.makedirs(REPORTS_DIR, exist_ok=True)

    # ── Public API ─────────────────────────────────────────────────────────────
    def record(self, test_id, module, test_name, status, duration,
               error_msg="", screenshot_path="", steps=""):
        self.results.append({
            "id":            test_id,
            "module":        module,
            "test_name":     test_name,
            "status":        status.upper(),   # PASS / FAIL / SKIP
            "duration":      round(duration, 2),
            "error_msg":     error_msg,
            "screenshot":    screenshot_path,
            "steps":         steps,
            "timestamp":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    def save(self):
        wb = Workbook()

        self._write_summary_sheet(wb)
        self._write_detail_sheet(wb)
        self._write_module_sheet(wb)
        self._write_failed_sheet(wb)
        self._write_charts_sheet(wb)

        # Remove the default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(EXCEL_REPORT)
        print(f"\n✅  Excel report saved → {EXCEL_REPORT}")
        return EXCEL_REPORT

    # ── Sheet builders ─────────────────────────────────────────────────────────
    def _write_summary_sheet(self, wb):
        ws = wb.create_sheet("📊 Summary", 0)
        ws.sheet_view.showGridLines = False

        total  = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASS")
        failed = sum(1 for r in self.results if r["status"] == "FAIL")
        skipped= sum(1 for r in self.results if r["status"] == "SKIP")
        pass_rate = (passed / total * 100) if total else 0
        end_time   = datetime.now()
        duration   = (end_time - self.start_time).total_seconds()

        # Title row
        ws.merge_cells("A1:H1")
        ws["A1"] = "🧠  CogniTest Android — Appium E2E Test Report"
        ws["A1"].font      = _title_font(16)
        ws["A1"].fill      = _fill(C_TITLE_BG)
        ws["A1"].alignment = _center()
        ws.row_dimensions[1].height = 40

        # Sub-title
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Generated: {end_time.strftime('%Y-%m-%d %H:%M:%S')}  |  Total Duration: {duration:.1f}s"
        ws["A2"].font      = Font(italic=True, color="AAAAAA", name="Calibri", size=10)
        ws["A2"].fill      = _fill(C_TITLE_BG)
        ws["A2"].alignment = _center()

        ws.append([])   # blank row

        # KPI cards
        kpi_headers = ["Total Tests", "Passed ✅", "Failed ❌", "Skipped ⚠️", "Pass Rate %", "Duration (s)"]
        kpi_values  = [total, passed, failed, skipped, f"{pass_rate:.1f}%", f"{duration:.1f}"]
        kpi_fills   = [C_HEADER_BG, C_PASS_FG, C_FAIL_FG, "7D5A00", "0D6B8E", "4A4A4A"]

        ws.append(kpi_headers)
        ws.append(kpi_values)

        for col_idx, (hdr, val, col) in enumerate(zip(kpi_headers, kpi_values, kpi_fills), start=1):
            hcell = ws.cell(row=4, column=col_idx, value=hdr)
            hcell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            hcell.fill      = _fill(col)
            hcell.alignment = _center()
            hcell.border    = _border()

            vcell = ws.cell(row=5, column=col_idx, value=val)
            vcell.font      = Font(bold=True, name="Calibri", size=14)
            vcell.alignment = _center()
            vcell.border    = _border()

        ws.row_dimensions[4].height = 28
        ws.row_dimensions[5].height = 36

        # Column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _write_detail_sheet(self, wb):
        ws = wb.create_sheet("📋 Test Details")
        ws.sheet_view.showGridLines = False

        headers = [
            "TC #", "Module", "Test Case Name", "Status",
            "Duration (s)", "Timestamp", "Steps", "Error / Notes"
        ]
        col_widths = [8, 22, 45, 10, 12, 20, 50, 55]

        # Header row
        ws.append(headers)
        for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value     = hdr
            cell.font      = _header_font()
            cell.fill      = _fill(C_HEADER_BG)
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # Data rows
        for row_num, r in enumerate(self.results, start=2):
            row_data = [
                r["id"], r["module"], r["test_name"], r["status"],
                r["duration"], r["timestamp"], r["steps"], r["error_msg"]
            ]
            ws.append(row_data)

            # Status colouring
            status = r["status"]
            if status == "PASS":
                bg, fg = C_PASS_BG, C_PASS_FG
            elif status == "FAIL":
                bg, fg = C_FAIL_BG, C_FAIL_FG
            else:
                bg, fg = C_SKIP_BG, C_SKIP_FG

            alt = C_ALT_ROW if row_num % 2 == 0 else "FFFFFF"
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = _border()
                cell.alignment = _center() if col_idx in (1, 4, 5, 6) else _left()
                cell.font      = _cell_font()
                if col_idx == 4:   # Status column
                    cell.font = Font(bold=True, color=fg, name="Calibri", size=10)
                    cell.fill = _fill(bg)
                else:
                    cell.fill = _fill(alt)
            ws.row_dimensions[row_num].height = 22

    def _write_module_sheet(self, wb):
        ws = wb.create_sheet("📦 Module Summary")
        ws.sheet_view.showGridLines = False

        # Aggregate by module
        modules = {}
        for r in self.results:
            m = r["module"]
            if m not in modules:
                modules[m] = {"total": 0, "pass": 0, "fail": 0, "skip": 0, "duration": 0.0}
            modules[m]["total"]    += 1
            modules[m]["duration"] += r["duration"]
            if r["status"] == "PASS":
                modules[m]["pass"] += 1
            elif r["status"] == "FAIL":
                modules[m]["fail"] += 1
            else:
                modules[m]["skip"] += 1

        headers = ["Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate %", "Total Duration (s)"]
        col_widths = [30, 10, 10, 10, 10, 14, 18]

        ws.append(headers)
        for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value     = hdr
            cell.font      = _header_font()
            cell.fill      = _fill(C_HEADER_BG)
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        for row_num, (module, stats) in enumerate(sorted(modules.items()), start=2):
            rate = (stats["pass"] / stats["total"] * 100) if stats["total"] else 0
            row_data = [
                module, stats["total"], stats["pass"], stats["fail"],
                stats["skip"], f"{rate:.1f}%", round(stats["duration"], 2)
            ]
            ws.append(row_data)
            alt = C_ALT_ROW if row_num % 2 == 0 else "FFFFFF"
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = _border()
                cell.alignment = _center() if col_idx > 1 else _left()
                cell.font      = _cell_font(bold=(col_idx == 1))
                cell.fill      = _fill(alt)
            ws.row_dimensions[row_num].height = 22

    def _write_failed_sheet(self, wb):
        ws = wb.create_sheet("❌ Failed Tests")
        ws.sheet_view.showGridLines = False

        failed = [r for r in self.results if r["status"] == "FAIL"]
        if not failed:
            ws["A1"] = "🎉 No failed test cases!"
            ws["A1"].font = Font(bold=True, color="1A6E1A", size=14)
            return

        headers = ["TC #", "Module", "Test Case Name", "Error Message", "Duration (s)", "Timestamp"]
        col_widths = [8, 22, 45, 70, 12, 20]

        ws.append(headers)
        for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value     = hdr
            cell.font      = _header_font()
            cell.fill      = _fill("8B0000")
            cell.alignment = _center()
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        for row_num, r in enumerate(failed, start=2):
            row_data = [r["id"], r["module"], r["test_name"], r["error_msg"], r["duration"], r["timestamp"]]
            ws.append(row_data)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = _border()
                cell.alignment = _center() if col_idx in (1, 5, 6) else _left()
                cell.font      = _cell_font()
                cell.fill      = _fill(C_FAIL_BG)
            ws.row_dimensions[row_num].height = 22

    def _write_charts_sheet(self, wb):
        ws = wb.create_sheet("📈 Charts")
        ws.sheet_view.showGridLines = False

        total  = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASS")
        failed = sum(1 for r in self.results if r["status"] == "FAIL")
        skipped= sum(1 for r in self.results if r["status"] == "SKIP")

        # Data table for charts
        ws["A1"] = "Status";    ws["B1"] = "Count"
        ws["A2"] = "Passed";    ws["B2"] = passed
        ws["A3"] = "Failed";    ws["B3"] = failed
        ws["A4"] = "Skipped";   ws["B4"] = skipped

        # Pie chart
        pie = PieChart()
        pie.title  = "Test Results Distribution"
        pie.width  = 15
        pie.height = 12
        data_ref   = Reference(ws, min_col=2, min_row=1, max_row=4)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=4)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        # Colour slices
        for idx, colour in enumerate(["00AA44", "DD2222", "EE9900"]):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = colour
            pie.series[0].dPt.append(pt)
        ws.add_chart(pie, "D1")

        # Bar chart — module pass rates
        module_ws = wb["📦 Module Summary"]
        if module_ws.max_row > 1:
            bar = BarChart()
            bar.type  = "col"
            bar.title = "Pass Rate by Module"
            bar.y_axis.title = "Pass Rate (%)"
            bar.x_axis.title = "Module"
            bar.width  = 25
            bar.height = 14
            bar.grouping = "clustered"
            data_ref2   = Reference(module_ws, min_col=3, min_row=1,
                                    max_col=5, max_row=module_ws.max_row)
            labels_ref2 = Reference(module_ws, min_col=1, min_row=2,
                                    max_row=module_ws.max_row)
            bar.add_data(data_ref2, titles_from_data=True)
            bar.set_categories(labels_ref2)
            ws.add_chart(bar, "D22")


# ── Singleton accessible from conftest ────────────────────────────────────────
reporter = ExcelReporter()
