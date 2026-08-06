"""
Generate SAST Vulnerability Excel Report from Bandit scan results.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_sast_excel_report(input_file="reports/sast_vulnerabilities.txt", output_file="reports/sast_vulnerability_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "🔒 Security SAST"
    ws.sheet_view.showGridLines = False

    c_header_bg = "1E3A5F"
    c_header_fg = "FFFFFF"
    c_border = "B0C4DE"

    def _border():
        side = Side(style="thin", color=c_border)
        return Border(left=side, right=side, top=side, bottom=side)

    def _fill(color):
        return PatternFill(fill_type="solid", fgColor=color)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "🔒  CogniTest SAST Vulnerability & Security Report"
    ws["A1"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=16)
    ws["A1"].fill = _fill("0D2B4E")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 36

    # Parse vulnerabilities
    issues = []
    total_issues = 0
    high_sev = 0
    medium_sev = 0
    low_sev = 0

    if os.path.exists(input_file):
        with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        current_issue = {}
        for line in content.splitlines():
            line_str = line.strip()
            if line_str.startswith(">> Issue:"):
                if current_issue:
                    issues.append(current_issue)
                current_issue = {"issue": line_str.replace(">> Issue:", "").strip(), "severity": "LOW", "confidence": "HIGH", "location": "N/A"}
                total_issues += 1
            elif line_str.startswith("Severity:"):
                parts = line_str.split("Confidence:")
                sev = parts[0].replace("Severity:", "").strip().upper()
                conf = parts[1].strip().upper() if len(parts) > 1 else "HIGH"
                current_issue["severity"] = sev
                current_issue["confidence"] = conf
                if "HIGH" in sev: high_sev += 1
                elif "MEDIUM" in sev: medium_sev += 1
                else: low_sev += 1
            elif line_str.startswith("Location:"):
                current_issue["location"] = line_str.replace("Location:", "").strip()

        if current_issue:
            issues.append(current_issue)

    # KPI Cards
    kpi_headers = ["Total Issues", "High Severity 🔴", "Medium Severity 🟡", "Low Severity 🟢", "Security Status"]
    status_str = "PASSED ✅" if high_sev == 0 else "ACTION REQUIRED ⚠️"
    kpi_values = [total_issues, high_sev, medium_sev, low_sev, status_str]
    kpi_colors = ["1E3A5F", "8B0000" if high_sev > 0 else "1A6E1A", "7D5A00", "1A6E1A", "1A6E1A" if high_sev == 0 else "8B0000"]

    ws.append([])
    ws.append(kpi_headers)
    ws.append(kpi_values)

    for col_idx, (hdr, val, col) in enumerate(zip(kpi_headers, kpi_values, kpi_colors), start=1):
        hcell = ws.cell(row=3, column=col_idx, value=hdr)
        hcell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        hcell.fill = _fill(col)
        hcell.alignment = _center()
        hcell.border = _border()

        vcell = ws.cell(row=4, column=col_idx, value=val)
        vcell.font = Font(bold=True, name="Calibri", size=13)
        vcell.alignment = _center()
        vcell.border = _border()

    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 32

    ws.append([])
    # Detail Table
    headers = ["#", "Severity", "Confidence", "Issue Description", "File Location"]
    col_widths = [8, 16, 16, 50, 45]

    ws.append(headers)
    header_row = 6
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=hdr)
        cell.font = Font(bold=True, color=c_header_fg, name="Calibri", size=11)
        cell.fill = _fill(c_header_bg)
        cell.alignment = _center()
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 26

    if not issues:
        ws.append([1, "LOW", "HIGH", "No SAST vulnerabilities detected in codebase.", "clean_scan"])
        for col_idx in range(1, 6):
            cell = ws.cell(row=7, column=col_idx)
            cell.border = _border()
            cell.alignment = _center()
    else:
        for idx, iss in enumerate(issues, start=1):
            ws.append([idx, iss.get("severity", "LOW"), iss.get("confidence", "HIGH"), iss.get("issue", ""), iss.get("location", "")])
            r_idx = header_row + idx
            for col_idx in range(1, 6):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.border = _border()
                cell.alignment = _center() if col_idx in (1, 2, 3) else _left()

    os.makedirs("reports", exist_ok=True)
    wb.save(output_file)
    print(f"✅ SAST Excel report generated → {output_file}")

if __name__ == "__main__":
    generate_sast_excel_report()
